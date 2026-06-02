#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Localizador de Leads sem Site no Google Places
Autor: Antigravity AI
Descrição: Encontra comércios locais que não possuem site cadastrado no Google Places
           para prospecção de serviços de Web Design.
"""

import os
import sys
import json
import requests
import pandas as pd
import urllib.parse

# Tenta importar dotenv para carregar variáveis de ambiente de um arquivo .env
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Garante compatibilidade de UTF-8 no terminal Windows para evitar erros com emojis
try:
    if sys.platform.startswith('win'):
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass


# Cores ANSI para o terminal (limpo e elegante)
class Cores:
    VERDE = '\033[92m'
    AMARELO = '\033[93m'
    VERMELHO = '\033[91m'
    AZUL = '\033[94m'
    NEGRITO = '\033[1m'
    RESET = '\033[0m'
    CIANO = '\033[96m'

def limpar_tela():
    os.system('cls' if os.name == 'nt' else 'clear')

def exibir_banner():
    limpar_tela()
    banner = f"""
{Cores.AZUL}{Cores.NEGRITO}======================================================================
     🔍 LOCALIZADOR DE LEADS SEM SITE - GOOGLE PLACES API 🔍
======================================================================{Cores.RESET}
 Desenvolvido para prospecção de serviços de Web Design de alto impacto.
    """
    print(banner)

def obter_validacao_whatsapp(telefone: str) -> str:
    """Valida se o telefone comercial é celular (WhatsApp) ou fixo."""
    if not telefone or telefone == "Não informado":
        return "Verificar"
    # Remove tudo que não for dígito
    nums = "".join(c for c in telefone if c.isdigit())
    
    # Se começar com 55 e tiver 12 ou 13 dígitos, remove o 55
    if nums.startswith("55") and len(nums) in [12, 13]:
        nums = nums[2:]
        
    # Celulares no Brasil têm 11 dígitos e o número local começa com 9: e.g. 71999998888
    # Telefones fixos têm 10 dígitos: e.g. 7133334444
    if len(nums) == 11 and nums[2] == '9':
        return "Sim"
    elif len(nums) == 10:
        if nums[2] == '9':
            return "Provável"
        elif nums[2] in ['2', '3', '4', '5']:
            return "Não"
        return "Verificar"
    else:
        # Caso tenha apenas 9 dígitos e comece com 9 (celular sem DDD)
        if len(nums) == 9 and nums[0] == '9':
            return "Sim"
        # Caso tenha 8 dígitos (fixo sem DDD)
        elif len(nums) == 8:
            if nums[0] in ['2', '3', '4', '5']:
                return "Não"
            return "Verificar"
        return "Verificar"

def obter_api_key():
    """Obtém a chave de API do Google Cloud de diferentes fontes possíveis."""
    # 1. Tenta carregar do ambiente (.env ou variável de sistema)
    api_key = os.environ.get("GOOGLE_API_KEY")
    if api_key:
        return api_key.strip()
        
    # 2. Tenta carregar de um arquivo local 'api_key.txt'
    if os.path.exists("api_key.txt"):
        try:
            with open("api_key.txt", "r", encoding="utf-8") as f:
                key = f.read().strip()
                if key:
                    return key
        except Exception:
            pass

    # 3. Caso não encontre, solicita ao usuário no terminal
    print(f"{Cores.AMARELO}[!] Chave de API do Google não encontrada no arquivo .env ou em variáveis de ambiente.{Cores.RESET}")
    print("Para usar este script, você precisa de uma API Key do Google Cloud com a API Places ativada.")
    print("Como obter: https://developers.google.com/maps/documentation/places/web-service/get-api-key\n")
    
    key_input = input(f"{Cores.NEGRITO}Digite sua API Key do Google Cloud (ou aperte Enter para sair): {Cores.RESET}").strip()
    if not key_input:
        print(f"\n{Cores.VERMELHO}[x] Execução cancelada pelo usuário. Chave de API não fornecida.{Cores.RESET}")
        sys.exit(0)
        
    # Pergunta se deseja salvar em um arquivo .env para as próximas vezes
    salvar = input(f"\nDeseja salvar esta chave no arquivo .env para não precisar digitar novamente? (s/n): ").strip().lower()
    if salvar in ['s', 'sim']:
        try:
            with open(".env", "a", encoding="utf-8") as f:
                f.write(f"\nGOOGLE_API_KEY={key_input}\n")
            print(f"{Cores.VERDE}[✓] Chave de API salva com sucesso no arquivo .env!{Cores.RESET}")
        except Exception as e:
            print(f"{Cores.AMARELO}[!] Não foi possível salvar o arquivo .env: {e}{Cores.RESET}")
            
    return key_input

def buscar_leads_nova_api(nicho, cidade, api_key):
    """
    Tenta buscar leads usando a nova API Google Places v1 (mais barata, rápida e moderna).
    Permite obter telefone e website em uma única requisição de texto.
    """
    url = "https://places.googleapis.com/v1/places:searchText"
    
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        # Solicita os campos necessários na máscara de campos
        "X-Goog-FieldMask": "places.id,places.displayName,places.formattedAddress,places.rating,places.userRatingCount,places.nationalPhoneNumber,places.websiteUri,places.googleMapsUri"
    }
    
    body = {
        "textQuery": f"{nicho} em {cidade}",
        "languageCode": "pt-BR"
    }
    
    print(f"\n{Cores.CIANO}[+] Consultando a API Google Places (New v1)...{Cores.RESET}")
    
    response = requests.post(url, headers=headers, json=body, timeout=15)
    
    # Tratamento de erro específico para a Nova API
    if response.status_code != 200:
        error_msg = response.text
        try:
            error_json = response.json()
            error_msg = error_json.get("error", {}).get("message", response.text)
        except Exception:
            pass
        
        # Se for um erro de permissão ou endpoint não habilitado, faremos o fallback
        raise RuntimeError(f"Erro na API v1 (Código {response.status_code}): {error_msg}")
        
    data = response.json()
    places = data.get("places", [])
    
    leads = []
    total_analisados = len(places)
    
    for place in places:
        # Verifica se o site existe
        website = place.get("websiteUri")
        
        # Filtragem: capturar apenas comércios sem site cadastrado
        if not website:
            nome = place.get("displayName", {}).get("text", "Nome não informado")
            telefone = place.get("nationalPhoneNumber", "Não informado")
            endereco = place.get("formattedAddress", "Não informado")
            rating = place.get("rating", 0.0)
            user_ratings = place.get("userRatingCount", 0)
            
            place_id = place.get("id")
            if place_id:
                maps_link = f"https://www.google.com/maps/search/?api=1&query=Google&query_place_id={place_id}"
                tipo_link_maps = "Link Direto"
            else:
                maps_link = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote_plus(nome + ' ' + endereco)}"
                tipo_link_maps = "Link de Pesquisa"
            
            leads.append({
                "Nome": nome,
                "Telefone": telefone,
                "Endereço": endereco,
                "Avaliação (Nota)": rating,
                "Avaliações (Total)": user_ratings,
                "Link do Google Maps": maps_link,
                "validacao_whatsapp": obter_validacao_whatsapp(telefone),
                "tipo_link_maps": tipo_link_maps
            })
            
    return leads, total_analisados

def buscar_leads_legacy_api(nicho, cidade, api_key):
    """
    Fallback usando a API Google Places antiga (Text Search + Place Details).
    Garante funcionamento mesmo se o usuário tiver apenas a API antiga habilitada no console.
    """
    print(f"\n{Cores.AMARELO}[!] Usando API Legacy (v0) como alternativa de fallback...{Cores.RESET}")
    
    # Passo 1: Text Search para encontrar estabelecimentos
    search_url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params = {
        "query": f"{nicho} em {cidade}",
        "key": api_key,
        "language": "pt-BR"
    }
    
    response = requests.get(search_url, params=params, timeout=15)
    
    if response.status_code != 200:
        raise RuntimeError(f"Erro no Text Search (Código {response.status_code}): {response.text}")
        
    data = response.json()
    
    if data.get("status") == "REQUEST_DENIED":
        error_msg = data.get("error_message", "Acesso negado pela API.")
        raise RuntimeError(f"Chave de API inválida ou sem permissão para Places API: {error_msg}")
        
    results = data.get("results", [])
    total_analisados = len(results)
    
    leads = []
    
    print(f"{Cores.CIANO}[+] Analisando detalhes de {total_analisados} estabelecimentos encontrados...{Cores.RESET}")
    
    for i, place in enumerate(results, start=1):
        place_id = place.get("place_id")
        nome = place.get("name")
        print(f"    [{i}/{total_analisados}] Verificando detalhes de: {nome[:30]}...", end="\r")
        
        if not place_id:
            continue
            
        # Passo 2: Detalhes do local para verificar telefone e website
        details_url = "https://maps.googleapis.com/maps/api/place/details/json"
        details_params = {
            "place_id": place_id,
            "fields": "name,formatted_phone_number,formatted_address,rating,user_ratings_total,website,url",
            "key": api_key,
            "language": "pt-BR"
        }
        
        try:
            details_resp = requests.get(details_url, params=details_params, timeout=10)
            if details_resp.status_code == 200:
                details_data = details_resp.json().get("result", {})
                
                website = details_data.get("website")
                # Filtra estabelecimentos sem site
                if not website:
                    nome_det = details_data.get("name", nome)
                    endereco_det = details_data.get("formatted_address", "Não informado")
                    
                    if place_id:
                        maps_link = f"https://www.google.com/maps/search/?api=1&query=Google&query_place_id={place_id}"
                        tipo_link_maps = "Link Direto"
                    else:
                        maps_link = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote_plus(nome_det + ' ' + endereco_det)}"
                        tipo_link_maps = "Link de Pesquisa"
                        
                    leads.append({
                        "Nome": nome_det,
                        "Telefone": details_data.get("formatted_phone_number", "Não informado"),
                        "Endereço": endereco_det,
                        "Avaliação (Nota)": details_data.get("rating", 0.0),
                        "Avaliações (Total)": details_data.get("user_ratings_total", 0),
                        "Link do Google Maps": maps_link,
                        "validacao_whatsapp": obter_validacao_whatsapp(details_data.get("formatted_phone_number", "Não informado")),
                        "tipo_link_maps": tipo_link_maps
                    })
        except Exception as e:
            # Continua verificando os outros mesmo se um falhar
            pass
            
    print(f"\n{Cores.VERDE}[✓] Análise de detalhes concluída.{Cores.RESET}")
    return leads, total_analisados

def formatar_e_salvar_excel(df, filename="leads_sem_site.xlsx"):
    """Salva os dados formatados em Excel com estilo profissional."""
    # Renomeia validacao_whatsapp para WhatsApp
    if "validacao_whatsapp" in df.columns:
        df = df.rename(columns={"validacao_whatsapp": "WhatsApp"})
    # Remove coluna tipo_link_maps se existir
    if "tipo_link_maps" in df.columns:
        df = df.drop(columns=["tipo_link_maps"])

    # Resolve arquivo bloqueado (ex: aberto no Excel) gerando um sufixo numérico
    actual_filename = filename
    if os.path.exists(filename):
        base, ext = os.path.splitext(filename)
        counter = 1
        while True:
            try:
                with open(actual_filename, "r+"):
                    pass
                break
            except FileNotFoundError:
                break
            except (IOError, PermissionError):
                actual_filename = f"{base}_{counter}{ext}"
                counter += 1
                if counter > 100:
                    break

    if actual_filename != filename:
        print(f"{Cores.AMARELO}[!] O arquivo '{filename}' está aberto ou bloqueado. Salvando como '{actual_filename}'...{Cores.RESET}")

    try:
        # Usa pandas com o openpyxl para formatar
        with pd.ExcelWriter(actual_filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Leads")
            
            # Formatação do Excel para um visual Premium
            workbook = writer.book
            worksheet = writer.sheets["Leads"]
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Estilos
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul Escuro Corporativo
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            border_thin = Side(border_style="thin", color="D3D3D3")
            border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
            
            # Formatando Cabeçalhos
            for col_num, column_title in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_cell
            
            # Formatando Células
            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border_cell
                    cell.font = Font(name="Calibri", size=11)
                    
                    # Centralizar Notas, Quantidades e WhatsApp
                    if col in [4, 5, 7]: # Colunas de Avaliação, Total de Avaliações e WhatsApp
                        cell.alignment = align_center
                    elif col == 6: # Coluna de Link do Google Maps
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                    else:
                        cell.alignment = align_left
                        
                    # Formata coluna do link do Maps como um link azul e sublinhado clicável no Excel
                    if col == 6 and str(cell.value).startswith("http"):
                        url_completa = cell.value
                        cell.value = f'=HYPERLINK("{url_completa}", "Abrir no Maps")'
                        cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
                        
            # Ajustando a largura das colunas dinamicamente
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        val_str = str(cell.value)
                        if val_str.startswith("=HYPERLINK"):
                            val_len = 14  # "Abrir no Maps" length
                        else:
                            val_len = len(val_str)
                        max_len = max(max_len, val_len)
                # Define largura com uma folga
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            # Força a largura da coluna F para 20
            worksheet.column_dimensions['F'].width = 20
                
            # Adiciona linhas de grade padrão
            worksheet.views.sheetView[0].showGridLines = True
            
        print(f"{Cores.VERDE}[✓] Leads salvos com sucesso e formatados no Excel: {Cores.NEGRITO}{filename}{Cores.RESET}")
        
    except Exception as e:
        # Fallback de segurança se falhar a formatação do Excel (ex: falta de openpyxl)
        print(f"\n{Cores.AMARELO}[!] Falha ao gerar arquivo Excel formatado ({e}).{Cores.RESET}")
        try:
            df.to_csv("leads_sem_site.csv", index=False, encoding="utf-8-sig")
            print(f"{Cores.VERDE}[✓] Arquivo de segurança salvo em CSV: {Cores.NEGRITO}leads_sem_site.csv{Cores.RESET}")
        except Exception as csv_error:
            print(f"{Cores.VERMELHO}[x] Falha crítica: não foi possível salvar em CSV: {csv_error}{Cores.RESET}")

def main():
    exibir_banner()
    api_key = obter_api_key()
    
    # Variáveis de entrada (pode ser interativo ou passado por argumento)
    nicho = ""
    cidade = ""
    
    # Se passar argumentos via linha de comando
    if len(sys.argv) > 2:
        nicho = sys.argv[1]
        cidade = sys.argv[2]
        print(f"{Cores.VERDE}[+] Entradas recebidas por linha de comando:{Cores.RESET}")
        print(f"    Nicho: {nicho}")
        print(f"    Cidade: {cidade}")
    else:
        # Entrada interativa por padrão
        print(f"{Cores.NEGRITO}Digite os filtros de busca para prospecção:{Cores.RESET}")
        nicho = input("➔ Nicho do comércio (ex: pet shop, barbearia, academia): ").strip()
        cidade = input("➔ Cidade e Estado (ex: Feira de Santana - BA): ").strip()
        
    if not nicho or not cidade:
        print(f"\n{Cores.VERMELHO}[x] Erro: O nicho e a cidade são obrigatórios!{Cores.RESET}")
        sys.exit(1)
        
    print(f"\n{Cores.VERDE}[+] Iniciando busca por '{nicho}' em '{cidade}'...{Cores.RESET}")
    
    leads = []
    total_analisados = 0
    
    try:
        # 1. Tenta usar a nova API Google Places v1
        try:
            leads, total_analisados = buscar_leads_nova_api(nicho, cidade, api_key)
        except Exception as api_v1_error:
            # Se der erro (ex: chave de API não tem acesso à API v1, ou erro de quota/permissão)
            # Mostra o aviso e tenta a API antiga como fallback
            print(f"{Cores.AMARELO}[!] Não foi possível usar a API do Places v1: {api_v1_error}{Cores.RESET}")
            leads, total_analisados = buscar_leads_legacy_api(nicho, cidade, api_key)
            
        # 2. Exibe os resultados obtidos
        print(f"\n{Cores.VERDE}======================================================================{Cores.RESET}")
        print(f"📊 {Cores.NEGRITO}RESULTADO DA BUSCA:{Cores.RESET}")
        print(f"➔ Estabelecimentos analisados: {total_analisados}")
        print(f"➔ Leads sem site encontrados: {Cores.VERDE}{Cores.NEGRITO}{len(leads)}{Cores.RESET}")
        print(f"{Cores.VERDE}======================================================================{Cores.RESET}")
        
        if len(leads) > 0:
            df = pd.DataFrame(leads)
            
            # Mostra prévia estilizada no terminal
            print(f"\n{Cores.NEGRITO}Prévia dos Leads sem Site encontrados:{Cores.RESET}")
            # Limita tamanho dos campos para caber no terminal e remove o link para exibição limpa
            df_preview = df.copy()
            if "Link do Google Maps" in df_preview.columns:
                df_preview = df_preview.drop(columns=["Link do Google Maps"])
            if "tipo_link_maps" in df_preview.columns:
                df_preview = df_preview.drop(columns=["tipo_link_maps"])
            if "validacao_whatsapp" in df_preview.columns:
                df_preview = df_preview.rename(columns={"validacao_whatsapp": "WhatsApp"})
            df_preview["Nome"] = df_preview["Nome"].apply(lambda x: x[:25] + "..." if len(x) > 25 else x)
            df_preview["Endereço"] = df_preview["Endereço"].apply(lambda x: x[:30] + "..." if len(x) > 30 else x)
            print(df_preview.to_string(index=False))
            print()
            
            # Salva o arquivo final excel/csv
            formatar_e_salvar_excel(df, "leads_sem_site.xlsx")
        else:
            print(f"\n{Cores.AMARELO}[!] Nenhum estabelecimento sem site foi encontrado para os critérios inseridos.{Cores.RESET}")
            
    except requests.exceptions.ConnectionError:
        print(f"\n{Cores.VERMELHO}[x] Erro de Conexão: Verifique sua internet.{Cores.RESET}")
    except requests.exceptions.Timeout:
        print(f"\n{Cores.VERMELHO}[x] Erro de Timeout: O servidor do Google demorou muito para responder.{Cores.RESET}")
    except Exception as e:
        print(f"\n{Cores.VERMELHO}[x] Ocorreu um erro inesperado: {e}{Cores.RESET}")
        
    print(f"\n{Cores.CIANO}➔ Script concluído com sucesso.{Cores.RESET}")

if __name__ == "__main__":
    main()
